#fsTimer - free, open source software for race timing.
#Copyright 2012-15 Ben Letham

#This program is free software: you can redistribute it and/or modify
#it under the terms of the GNU General Public License as published by
#the Free Software Foundation, either version 3 of the License, or
#(at your option) any later version.

#This program is distributed in the hope that it will be useful,
#but WITHOUT ANY WARRANTY; without even the implied warranty of
#MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#GNU General Public License for more details.

#You should have received a copy of the GNU General Public License
#along with this program.  If not, see <http://www.gnu.org/licenses/>.

#The author/copyright holder can be contacted at bletham@gmail.com
'''Handling of the window dedicated to importation of pre-registration'''

import gi
gi.require_version('Gtk', '3.0')
from gi.repository import Gtk
from enum import Enum
import fstimer.gui
import os, csv, json
import datetime
import openpyxl
from fstimer.gui.util_classes import GtkStockButton
from openpyxl import load_workbook

class ImportFormat(Enum):
     CVS = 1
     EXCEL = 2

class ComboValueError(Exception):
    '''Exception launched when decoding reveals an invalid value for a combo field'''
    pass

class ImportPreRegWin(Gtk.Window):
    '''Handling of the window dedicated to importation of pre-registration'''

    def __init__(self, path, fields, fieldsdic):
        '''Builds and display the importation window'''
        super(ImportPreRegWin, self).__init__(Gtk.WindowType.TOPLEVEL)
        self.path = path
        self.fields = fields
        self.fieldsdic = fieldsdic
        self.modify_bg(Gtk.StateType.NORMAL, fstimer.gui.bgcolor)
        fname = os.path.abspath(
            os.path.join(
                os.path.dirname(os.path.abspath(__file__)),
                '../data/icon.png'))
        self.set_icon_from_file(fname)
        self.set_title('fsTimer - ' + os.path.basename(path))
        self.set_position(Gtk.WindowPosition.CENTER)
        self.connect('delete_event', lambda b, jnk: self.hide())
        self.set_border_width(10)
        self.set_size_request(600, 400)
        # Start with some intro text.
        label1 = Gtk.Label(
            label=(_('Select a pre-registration csv or Excel file to import.\n\n'
                   'If the spreadsheet has accented or non-Latin characters, it must be in utf-8,\n'
                   'see Section 2.2 of the documentation for instructions.')))
        #Continue to the load file.
        btnFILE = GtkStockButton('open',_("Open"))
        ## Textbuffer
        textbuffer = Gtk.TextBuffer()
        try:
            textbuffer.create_tag("blue", foreground="blue")
            textbuffer.create_tag("red", foreground="red")
        except TypeError:
            pass
        textview = Gtk.TextView()
        textview.set_buffer(textbuffer)
        textview.set_editable(False)
        textview.set_cursor_visible(False)
        textsw = Gtk.ScrolledWindow()
        textsw.set_shadow_type(Gtk.ShadowType.ETCHED_IN)
        textsw.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
        textsw.add(textview)
        textalgn = Gtk.Alignment.new(0, 0, 1, 1)
        textalgn.add(textsw)
        hbox2 = Gtk.HBox(False, 5)
        btnFILE.connect('clicked', self.select_preregistration, textbuffer)
        btn_algn = Gtk.Alignment.new(1, 0, 1, 0)
        hbox2.pack_start(btnFILE, False, False, 0)
        hbox2.pack_start(btn_algn, True, True, 0)
        ## buttons
        btnOK = GtkStockButton('ok',_("OK"))
        btnOK.connect('clicked', lambda b: self.hide())
        cancel_algn = Gtk.Alignment.new(0, 0, 1, 0)
        hbox3 = Gtk.HBox(False, 10)
        hbox3.pack_start(cancel_algn, True, True, 0)
        hbox3.pack_start(btnOK, False, False, 0)
        vbox = Gtk.VBox(False, 0)
        vbox.pack_start(label1, False, False, 5)
        vbox.pack_start(hbox2, False, True, 5)
        vbox.pack_start(textalgn, True, True, 5)
        vbox.pack_start(hbox3, False, False, 0)
        self.add(vbox)
        self.show_all()

    def propose_advanced_import(self, csv_fields, textbuffer1, importFormat):
        '''Propose advanced import mechanism where project fields can be build
           from the csv ones using python expressions'''
        self.advancedwin = Gtk.Window(Gtk.WindowType.TOPLEVEL)
        self.advancedwin.modify_bg(Gtk.StateType.NORMAL, fstimer.gui.bgcolor)
        self.advancedwin.set_transient_for(self)
        self.advancedwin.set_modal(True)
        if (importFormat==ImportFormat.EXCEL):
            self.advancedwin.set_title(_('fsTimer - EXCEL import'))
        else:
            self.advancedwin.set_title(_('fsTimer - CSV import'))
        self.advancedwin.set_position(Gtk.WindowPosition.CENTER)
        self.advancedwin.set_border_width(20)
        self.advancedwin.set_size_request(600, 400)
        self.advancedwin.connect('delete_event', lambda b, jnk_unused: self.advancedwin.hide())
        # top label
        if (importFormat==ImportFormat.EXCEL):
            toplabel = Gtk.Label(_("For each field, specify the corresponding EXCEL column.\n"))
        else:
            toplabel = Gtk.Label(_("For each field, specify the corresponding CSV column.\n"))
        # Treeview with 3 columns : field, combobox and free text
        self.fieldview = Gtk.TreeView()
        self.fieldview.set_grid_lines(Gtk.TreeViewGridLines.BOTH)
        # Associated model with 5 columns : the 4th one is a boolean indicating
        # whether the 3rd one (Advanced mapping) should be sensitive
        self.fieldsmodel = Gtk.ListStore(str, str, str, bool)
        for field in self.fields:
            self.fieldsmodel.append([field, field if field in csv_fields else '-- select --', '', False])
        self.fieldview.set_model(self.fieldsmodel)
        # build first column (project field)
        column = Gtk.TreeViewColumn('Field', Gtk.CellRendererText(), text=0)
        self.fieldview.append_column(column)
        # vuild 2nd column (csv field to be used, as a combo box)
        combo_renderer = Gtk.CellRendererCombo()
        liststore_csv_fields = Gtk.ListStore(str)
        liststore_csv_fields.append(['-- Leave empty --'])
        for field in csv_fields:
            liststore_csv_fields.append([field])
        liststore_csv_fields.append(['-- Advanced expression --'])
        combo_renderer.set_property("model", liststore_csv_fields)
        combo_renderer.set_property("text-column", 0)
        combo_renderer.set_property("editable", True)
        combo_renderer.set_property("has-entry", False)
        if (importFormat==ImportFormat.EXCEL):
            column = Gtk.TreeViewColumn('EXCEL column', combo_renderer, text=1)
        else:
            column = Gtk.TreeViewColumn('CSV column', combo_renderer, text=1)
        self.fieldview.append_column(column)
        # build the 3rd column (Advanced mapping)
        advanced_renderer = Gtk.CellRendererText()
        column = Gtk.TreeViewColumn('Advanced mapping', advanced_renderer, text=2, sensitive=3, editable=3)
        self.fieldview.append_column(column)
        # handler for the combo changes
        combo_renderer.connect("edited", self.combo_changed)
        advanced_renderer.connect("edited", self.text_changed)
        # And put it in a scrolled window, in an alignment
        fieldsw = Gtk.ScrolledWindow()
        fieldsw.set_shadow_type(Gtk.ShadowType.ETCHED_IN)
        fieldsw.set_policy(Gtk.PolicyType.AUTOMATIC, Gtk.PolicyType.AUTOMATIC)
        fieldsw.add(self.fieldview)
        fieldalgn = Gtk.Alignment.new(0, 0, 1, 1)
        fieldalgn.add(fieldsw)
        # a text buffer for errors
        textbuffer2 = Gtk.TextBuffer()
        try:
            textbuffer2.create_tag("red", foreground="red")
            textbuffer2.create_tag("blue", foreground="blue")
        except TypeError:
            pass
        textview = Gtk.TextView()
        textview.set_buffer(textbuffer2)
        textview.set_editable(False)
        textview.set_cursor_visible(False)
        # hbox for the buttons
        hbox = Gtk.HBox(False, 0)
        btnCANCEL = GtkStockButton('close',"Cancel")
        btnCANCEL.connect('clicked', self.advanced_import_cancel, textbuffer1, importFormat)
        btnOK = GtkStockButton('ok',"OK")
        btnOK.connect('clicked', self.advanced_import_ok, textbuffer1, textbuffer2, importFormat)
        alignOK = Gtk.Alignment.new(1, 0, 0, 0)
        alignOK.add(btnOK)
        hbox.pack_start(btnCANCEL, False, True, 0)
        hbox.pack_start(alignOK, True, True, 0)
        # populate
        vbox = Gtk.VBox(False, 10)
        vbox.pack_start(toplabel, False, False, 0)
        vbox.pack_start(fieldalgn, True, True, 0)
        vbox.pack_start(textview, False, False, 0)
        vbox.pack_start(hbox, False, False, 10)
        self.advancedwin.add(vbox)
        self.advancedwin.show_all()

    def advanced_import_cancel(self, widget_unused, textbuffer, importFormat):
        '''If cancel is pressed in the advanced import window'''
        self.advancedwin.hide()
        iter_end = textbuffer.get_end_iter()
        textbuffer.insert_with_tags_by_name(iter_end, 'Nothing done.', 'blue')

    def combo_changed(self, widget_unused, gtkpath, text):
        '''Handles a change in the combo boxes' selections'''
        self.fieldsmodel[gtkpath][1] = text
        if text == '-- Advanced expression --':
            if len(self.fieldsmodel[gtkpath][2]) == 0:
                self.fieldsmodel[gtkpath][2] = '-- enter python expression --'
            self.fieldsmodel[gtkpath][3] = True
        else:
            self.fieldsmodel[gtkpath][3] = False

    def text_changed(self, widget_unused, gtkpath, text):
        '''Handles a change in the advanced boxes'''
        self.fieldsmodel[gtkpath][2] = text

    def advanced_import_ok(self, jnk_unused, textbuffer1, textbuffer2, importFormat):
        '''Handles click on OK button in the advanced interface'''
        textbuffer2.delete(textbuffer2.get_start_iter(), textbuffer2.get_end_iter())
        self.fields_mapping = {}
        for gtkpath in range(len(self.fieldsmodel)):
            field = self.fieldsmodel[gtkpath][0]
            csv_col = self.fieldsmodel[gtkpath][1]
            if csv_col == '-- select --':
                textbuffer2.insert_with_tags_by_name(textbuffer2.get_end_iter(), 'Nothing selected for field %s' % field, 'red')
                return
            elif csv_col == '-- Leave empty --' or csv_col == None:
                self.fields_mapping[field] = lambda reg, col=csv_col: ''
            elif csv_col == '-- Advanced expression --':
                try:
                    code = compile(self.fieldsmodel[gtkpath][2], '', 'eval')
                    self.fields_mapping[field] = lambda reg, code=code: eval(code)
                except SyntaxError as e:
                    iter_end = textbuffer2.get_end_iter()
                    textbuffer2.insert_with_tags_by_name(iter_end, 'Invalid syntax for expression of field %s: ' % field, 'red')
                    textbuffer2.insert_with_tags_by_name(textbuffer2.get_end_iter(), str(e), 'blue')
                    return
            else:
                self.fields_mapping[field] = lambda reg, col=csv_col: reg[col]
        self.advancedwin.hide()
        self.import_data(textbuffer1, importFormat)

    def build_fields_mapping(self, import_fields, textbuffer, importFormat):
        '''Maps cvs fields to project fields and creates a dictionnary
           of lambdas to apply to a csv entry to extract each field.
           Some entries may contain strings instead of lambdas, meaning
           that the project column's value is equal to that csv column's value'''
        iter_end = textbuffer.get_end_iter()
        fields_use = [field for field in import_fields if field in self.fields]
        if (importFormat==ImportFormat.CVS):
            textbuffer.insert_with_tags_by_name(iter_end, 'Matched csv fields: ', 'blue')
        else:
            textbuffer.insert_with_tags_by_name(iter_end, 'Matched excel fields: ', 'blue')
        textbuffer.insert(iter_end, ', '.join(fields_use) + '\n')
        fields_ignore = [field for field in import_fields if field not in self.fields]
        if (importFormat==ImportFormat.CVS):
            textbuffer.insert_with_tags_by_name(iter_end, 'Did not match csv fields: ', 'red')
        else:
            textbuffer.insert_with_tags_by_name(iter_end, 'Did not match excel fields: ', 'red')
        textbuffer.insert(iter_end, ', '.join(fields_ignore) + '\n')
        fields_notuse = [field for field in self.fields if field not in import_fields]
        if (importFormat==ImportFormat.CVS):
            textbuffer.insert_with_tags_by_name(iter_end, 'Did not find in csv: ', 'red')
        else:
            textbuffer.insert_with_tags_by_name(iter_end, 'Did not find in excel: ', 'red')
        textbuffer.insert(iter_end, ', '.join(fields_notuse) + '\n')
        self.propose_advanced_import(import_fields, textbuffer, importFormat)

    def select_preregistration(self, jnk_unused, textbuffer):
        '''Handle selection of a pre-reg file using a filechooser'''
        chooser = Gtk.FileChooserDialog(title='Select pre-registration csv', parent=self, action=Gtk.FileChooserAction.OPEN, buttons=('Cancel', Gtk.ResponseType.CANCEL, 'OK', Gtk.ResponseType.OK))
        ffilterCSV = Gtk.FileFilter()
        ffilterCSV.set_name('csv files')
        ffilterCSV.add_pattern('*.csv')
        chooser.add_filter(ffilterCSV)
        ffilterExcel = Gtk.FileFilter()
        ffilterExcel.set_name('excel files')
        ffilterExcel.add_pattern('*.xls')
        ffilterExcel.add_pattern('*.xlsx')
        chooser.add_filter(ffilterExcel)
        chooser.set_current_folder(self.path)
        response = chooser.run()
        if response == Gtk.ResponseType.OK:
            filename = chooser.get_filename()
            textbuffer.delete(textbuffer.get_start_iter(), textbuffer.get_end_iter())
            textbuffer.set_text('Loading '+os.path.basename(filename)+'...\n')
            if (filename.endswith(".xls") or filename.endswith(".xlsx")):
                print("Reading Exel from: " + filename)
                self.csvreg = []
                self.wb = load_workbook(filename=filename, read_only=True)
                self.ws = self.wb[self.wb.sheetnames[0]]
                excel_fields = []
                rowCount = 0
                for row in self.ws.rows:
                    if (rowCount == 0):
                        for cell in row:
                            excel_fields.append(cell.value)
                    else:
                        colCount = 0
                        rowDict = {}
                        for cell in row:
                            value = cell.value
                            key = excel_fields[colCount]
                            rowDict[key] = value
                            colCount = colCount + 1
                        self.csvreg.append(rowDict)
                    rowCount = rowCount + 1
                iter_end = textbuffer.get_end_iter()
                textbuffer.insert_with_tags_by_name(iter_end, 'Found excel fields: ', 'blue')
                textbuffer.insert(iter_end, ', '.join(excel_fields) + '\n')
                self.build_fields_mapping(excel_fields, textbuffer, ImportFormat.EXCEL)
            else:
                print("Reading csv from: " + filename)
                try:
                    fin = csv.DictReader(open(filename, 'r', encoding='utf-8'))
                    self.csvreg = []
                    for row in fin:
                        self.csvreg.append(row)
                    csv_fields = self.csvreg[0].keys()
                    iter_end = textbuffer.get_end_iter()
                    textbuffer.insert_with_tags_by_name(iter_end, 'Found csv fields: ', 'blue')
                    textbuffer.insert(iter_end, ', '.join(csv_fields) + '\n')
                    self.build_fields_mapping(csv_fields, textbuffer, ImportFormat.CVS)
                except (IOError, IndexError, csv.Error):
                    iter_end = textbuffer.get_end_iter()
                    textbuffer.insert_with_tags_by_name(iter_end, 'Error! Could not open file, or no data found in file. Nothing was imported, try again.', 'red')
        chooser.destroy()

    def import_data(self, textbuffer, importFormat):
        '''Implements the actual import of the csv data'''
        textbuffer.insert(textbuffer.get_end_iter(), 'Importing registration data...\n')
        preregdata = []
        row = 1
        for reg in self.csvreg:
            #print("csv type of reg: " + str(type(reg)))
            #for k in reg:
            #    print("Key: " + str(k) + " value: " + str(reg[k]))
            tmpdict = {}
            for field in self.fields:
                value = self.fields_mapping[field](reg)
                if value and self.fieldsdic[field]['type'] == 'combobox':
                    if value not in self.fieldsdic[field]['options']:
                        optstr = '"' + '", "'.join(self.fieldsdic[field]['options']) + '", and blank'
                        errstr = """Error in csv row %d!
Found value "%s" in field "%s". Not a valid value!
Valid values (case sensitive) are: %s.
Correct the error and try again.""" % (row+1, value, field, optstr)
                        textbuffer.insert_with_tags_by_name(textbuffer.get_end_iter(), errstr, 'red')
                        return
                tmpdict[field] = value
            preregdata.append(tmpdict.copy())
            row += 1
        with open(os.path.join(self.path, os.path.basename(self.path)+'_registration_prereg.json'), 'w', encoding='utf-8') as fout:
            json.dump(preregdata, fout)
        textbuffer.insert_with_tags_by_name(textbuffer.get_end_iter(), 'Imported ' + str(row) + ' records\n', 'blue')
        textbuffer.insert_with_tags_by_name(textbuffer.get_end_iter(), 'Success! Imported pre-registration saved to '+os.path.basename(self.path)+'_registration_prereg.json\nFinished!', 'blue')
